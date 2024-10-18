#   ************************************************************************************************************************* 

start_time = '18:00:00'
end_time = '20:00:00'

# Read python dates.txt

import re

# Open the file and read the contents
with open('python dates.txt', 'r') as file:
    content = file.read()

# Use regex to extract the list of classes taken dates
match = re.search(r'classes_taken_dates\s*=\s*(\[[^\]]*\])', content)
if match:
    # Evaluate the extracted string as a Python list
    class_dates = eval(match.group(1))
    # print("class_dates:", class_dates)
else:
    print("No 'classes_taken_dates' found in the file.")


# Use regex to extract the list of missed classes taken
match = re.search(r'classes_missed_dates\s*=\s*(\[[^\]]*\])', content)
if match:
    # Evaluate the extracted string as a Python list
    miss_dates = eval(match.group(1))
    # print("miss_dates:", miss_dates)
else:
    print("No 'classes_taken_dates' found in the file.")

# Use regex to extract the list of exam dates
match = re.search(r'exams_dates\s*=\s*(\[[^\]]*\])', content)
if match:
    # Evaluate the extracted string as a Python list
    exam_dates = eval(match.group(1))
    # print("exam_dates:", exam_dates)
else:
    print("No 'classes_taken_dates' found in the file.")

# print(type(class_list))
# ***********************************************


# Read stud_list.txt

stud_list = []
with open("stud_list.txt") as f:
    data = f.readline()

    while( data != '') :
        # print(data)
        stud_list.append(data.strip())
        data = f.readline()

# print(stud_list)
# ***********************************************

# Create object with properties as name and dates

desired_fieldnames = []
class MyObject:
    def __init__(self, name, class_dates, miss_dates, exam_dates):
        self.name = name  # Store the name as a property

        
        # Handle properties from list1
        for date in class_dates:
            attr_name = date.replace('/', '_')
            # print(attr_name)
            setattr(self, attr_name, 0)  # Set default value to 0

        # Handle properties from list2
        for value in miss_dates:
            attr_name = value.replace('/', '_')
            # print(attr_name)
            setattr(self, attr_name, 0)  # Set default value to 1 (or another default)

        # Handle properties from list3
        for item in exam_dates:
            attr_name = item.replace('/', '_')
            # print(attr_name)
            setattr(self, attr_name, 0)  # Set default value to 2 (or another default)

        self.max_allowed_attendance = 2 * len(class_dates)
        self.total_marked = 0    # Store count of total marked attendance
        self.proxy_ct = 0        # Proxy includes invalid_dates, (> 2) marked and invalid timing
        self.non_teaching_dates = []  # Store the count of non-teaching days
        self.invalid_timing = [] # Store the count of invalid timings on teaching days



desired_fieldnames.append('name')

for date in class_dates:
    attr_name = date.replace('/', '_')
    desired_fieldnames.append(attr_name)

for value in miss_dates:
    attr_name = value.replace('/', '_')
    desired_fieldnames.append(attr_name)

for item in exam_dates:
    attr_name = item.replace('/', '_')
    desired_fieldnames.append(attr_name)

desired_fieldnames.append('max_allowed_attendance')
desired_fieldnames.append('total_marked')
desired_fieldnames.append('proxy_ct')
desired_fieldnames.append('non_teaching_dates')
desired_fieldnames.append('invalid_timing')

# Generate obj list with names as student list

obj_list = []
stud_map = {}

idx = 0

for name in stud_list :
    obj = MyObject(name, class_dates, miss_dates, exam_dates)
    obj_list.append(obj)

    stud_map[name] = idx
    idx += 1

# for x in obj_list :
#     print(x.name)

# print(f'Total students = {len(obj_list)}')

# ***************************************************************



# Read input attendance & store each row in 'input_data' list

import pandas as pd

# Read the CSV file
df = pd.read_csv('input_attendance.csv')

# Convert each row to a list and store in a list of lists
input_data = df.values.tolist()

# ***********************************************





# Update attendance

from datetime import datetime

for data in input_data :

    date, time = data[0].split()
    identity = data[1]

    time_obj = datetime.strptime(time, "%H:%M:%S").time()
    start_time_obj = datetime.strptime(start_time, "%H:%M:%S").time()
    end_time_obj = datetime.strptime(end_time, "%H:%M:%S").time()

    if(identity in stud_list) :

        obj_list_idx = stud_map[identity]

        # Update total marked
        current_value = getattr(obj_list[obj_list_idx], 'total_marked', -1)
        setattr(obj_list[obj_list_idx], 'total_marked', current_value + 1)

        if(date in class_dates) :
            
            if (start_time_obj <= time_obj <= end_time_obj) :
                
                dynamic_attr_name = date.replace('/', '_')
                current_value = getattr(obj_list[obj_list_idx], dynamic_attr_name, -1)

                if(current_value >= 2) :

                    # Proxy update
                    current_value2 = getattr(obj_list[obj_list_idx], 'proxy_ct', -1)
                    setattr(obj_list[obj_list_idx], 'proxy_ct', current_value2 + 1)

                setattr(obj_list[obj_list_idx], dynamic_attr_name, current_value + 1)
            else :

                # invalid timing
                date_modified = date.replace('/', '_')
                obj_list[obj_list_idx].invalid_timing.append(data[0])

                # Proxy update
                current_value = getattr(obj_list[obj_list_idx], 'proxy_ct', -1)
                setattr(obj_list[obj_list_idx], 'proxy_ct', current_value + 1)
        else :

            # Non-teching days
            date_modified = date.replace('/', '_')
            obj_list[obj_list_idx].non_teaching_dates.append(date_modified)

            # Proxy update
            current_value = getattr(obj_list[obj_list_idx], 'proxy_ct', -1)
            setattr(obj_list[obj_list_idx], 'proxy_ct', current_value + 1)


# print('ALL DONE!')


#   *******************************************************************************************

# print(desired_fieldnames)

import csv

# Check how many entries are in obj_list before writing
# print(f'Number of entries in obj_list: {len(obj_list)}')

# Open a CSV file to write the generated data
with open('output.csv', mode='w', newline='') as file:
    # Initialize the DictWriter with desired fieldnames
    writer = csv.DictWriter(file, fieldnames=desired_fieldnames)

    # Write the header (fieldnames) only once
    writer.writeheader()

    for item in obj_list:
        curr_obj = item
        
        # Print the name or identifier of the current object
        # print(curr_obj.name)

        curr_obj_dict = {}

        # Build a dictionary from the object's properties
        for prop in desired_fieldnames:
            # print(prop)
            curr_obj_dict[prop] = getattr(curr_obj, prop)

            if(prop == 'invalid_timing') :

                if(curr_obj_dict[prop] == []) :
                    curr_obj_dict[prop] = 'NA'
                
            if(prop == 'non_teaching_dates') :
                
                if(curr_obj_dict[prop] == []) :
                    curr_obj_dict[prop] = 'NA'

        # Debugging: print current object's property dictionary
        # print(f'Writing row: {curr_obj_dict}')  # Debugging step
        
        # Write the current object's property values as a dictionary
        writer.writerow(curr_obj_dict)

#   *******************************************************************************************

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load the CSV file
file_path = 'output.csv'  # Input file path (your uploaded file)
df = pd.read_csv(file_path)

# Define dark color fills
dark_red_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")  # Dark Red
dark_green_fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")  # Dark Green
dark_yellow_fill = PatternFill(start_color="B8860B", end_color="B8860B", fill_type="solid")  # Dark Yellow

# Save the dataframe to an Excel file to apply the styling
output_file_path = 'output_styled.xlsx'  # Output file path
df.to_excel(output_file_path, index=False, engine='openpyxl')  

# Load the workbook and worksheet
wb = load_workbook(output_file_path)
ws = wb.active

# Apply conditional formatting to columns B to J (which correspond to columns 2 to 10 in 1-indexed Excel)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=10):  # Only columns B to J
    for cell in row:
        if cell.value > 2:
            cell.fill = dark_red_fill
        elif cell.value == 2:
            cell.fill = dark_green_fill
        elif cell.value == 1:
            cell.fill = dark_yellow_fill

# Save the styled Excel file
wb.save(output_file_path)